#!/usr/bin/env python3
"""
Instagram Publisher — DEVOPSRAIZ
================================

Lê o calendário editorial (XLSX) e publica / agenda posts no Instagram
usando a Instagram Graph API oficial da Meta.

Autor: Trilha DEVOPSRAIZ
Canal: @devopsraiz_oficial
Data-alvo: lançamento 21/04/2026

-------------------------------------------------------------------------------
PRÉ-REQUISITOS (uma vez só)
-------------------------------------------------------------------------------
1. Sua conta @devopsraiz_oficial precisa ser PROFISSIONAL (Business/Creator).
   No app Instagram: Menu > Configurações > Conta > Mudar para conta profissional.

2. Conecte sua conta Instagram a uma Página do Facebook:
   https://www.facebook.com/business/help/898752960195806

3. Crie um app em https://developers.facebook.com:
   - Tipo: Business
   - Adicione o produto "Instagram Graph API"
   - Permissões que você vai pedir no review:
       * instagram_basic
       * instagram_content_publish
       * pages_show_list
       * pages_read_engagement

4. Gere um token de LONGA duração (60 dias, renovável):
   https://developers.facebook.com/tools/explorer/
   Troque o short-lived token por long-lived em:
   https://graph.facebook.com/v21.0/oauth/access_token
     ?grant_type=fb_exchange_token
     &client_id={APP_ID}
     &client_secret={APP_SECRET}
     &fb_exchange_token={SHORT_LIVED_TOKEN}

5. Descubra o ID da sua conta Instagram Business:
   GET https://graph.facebook.com/v21.0/me/accounts?access_token=TOKEN
   → pegue o ID da Página → GET /{PAGE_ID}?fields=instagram_business_account

6. As imagens precisam estar HOSPEDADAS em URL público (Meta baixa do URL).
   Opções rápidas:
     - Cloudflare R2 / AWS S3 com CDN (+ cheap)
     - GitHub raw file (gratuito, OK pra começar)
     - Cloudinary free tier
   Configure a base URL em MEDIA_BASE_URL abaixo.

-------------------------------------------------------------------------------
.env esperado (coloque no mesmo diretório):
-------------------------------------------------------------------------------
IG_USER_ID=178414xxxxxx               # ID da conta Instagram Business
META_ACCESS_TOKEN=EAAJx...            # token long-lived (60 dias)
MEDIA_BASE_URL=https://cdn.devopsraiz.com.br/posts   # onde estão as PNGs
CALENDAR_XLSX=../calendario-editorial-30-dias.xlsx
CREATIVES_DIR=../criativos

-------------------------------------------------------------------------------
USO
-------------------------------------------------------------------------------
    python instagram_publisher.py plan              # mostra próximos 7 posts
    python instagram_publisher.py publish --day 1   # publica post do dia 1
    python instagram_publisher.py publish --today   # publica post do dia atual
    python instagram_publisher.py dry-run --day 3   # simula sem chamar API
    python instagram_publisher.py check             # valida token e IG ID

Recomendação de agendamento via cron (Linux/macOS):
    # todo dia 08:00 publica o post do dia
    0 8 * * * cd /caminho/automacao && /usr/bin/python3 instagram_publisher.py publish --today

No Windows, use Agendador de Tarefas (Task Scheduler) apontando pro mesmo comando.
"""

import argparse
import json
import os
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import requests
from openpyxl import load_workbook


# -------------------------------------------------------------------- config --
def load_env(env_path: str = ".env") -> dict:
    """Carrega variáveis do .env sem depender de python-dotenv."""
    config = {}
    p = Path(env_path)
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            config[k.strip()] = v.strip().strip('"').strip("'")
    # Env vars sobrescrevem .env
    for key in ["IG_USER_ID", "META_ACCESS_TOKEN", "MEDIA_BASE_URL",
                "CALENDAR_XLSX", "CREATIVES_DIR"]:
        if os.environ.get(key):
            config[key] = os.environ[key]
    return config


# -------------------------------------------------------------- xlsx loader --
def load_calendar(xlsx_path: str) -> list[dict]:
    """Lê a aba 'Calendario 30 dias' e devolve uma lista de posts."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Calendario 30 dias"]
    headers = [c.value for c in ws[1]]
    posts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        posts.append(dict(zip(headers, row)))
    return posts


# ------------------------------------------------------ image file resolver --
def creative_filename_for(post: dict) -> Optional[str]:
    """Mapeia o dia do post para um arquivo PNG em /criativos."""
    dia = int(post["Dia"])
    # Mapeamento dia -> arquivo de criativo correspondente
    mapping = {
        1: "01-apresentacao-capa.png",
        2: "02-docker-vm-vs-container.png",
        3: "11-salario-devops-2026.png",
        4: "03-aws-custos-queimando.png",
        5: "07-rag-ia-pipeline.png",
        8: "04-owasp-top-10.png",
        10: "05-saas-multi-tenant.png",
        15: "06-slo-uptime-table.png",
        18: "08-frase-junior-vs-senior.png",
        30: "09-fechamento-trilha-completa.png",
    }
    return mapping.get(dia)


# ---------------------------------------------------- instagram graph api --
GRAPH = "https://graph.instagram.com/v21.0"


class IGClient:
    def __init__(self, ig_user_id: str, access_token: str):
        self.ig_id = ig_user_id
        self.token = access_token

    def _post(self, path: str, params: dict):
        url = f"{GRAPH}/{path}"
        params = {**params, "access_token": self.token}
        r = requests.post(url, data=params, timeout=60)
        if r.status_code >= 400:
            raise RuntimeError(f"Meta API error {r.status_code}: {r.text}")
        return r.json()

    def _get(self, path: str, params: dict | None = None):
        url = f"{GRAPH}/{path}"
        params = {**(params or {}), "access_token": self.token}
        r = requests.get(url, params=params, timeout=60)
        if r.status_code >= 400:
            raise RuntimeError(f"Meta API error {r.status_code}: {r.text}")
        return r.json()

    def check_identity(self) -> dict:
        """Valida se o token tem acesso à conta Instagram."""
        return self._get("me", {"fields": "id,username,account_type,followers_count,media_count"})

    def refresh_token(self) -> dict:
        """Renova o token long-lived (válido por mais 60 dias).
        Endpoint específico do Instagram Graph API."""
        url = "https://graph.instagram.com/refresh_access_token"
        r = requests.get(url, params={
            "grant_type": "ig_refresh_token",
            "access_token": self.token,
        }, timeout=60)
        if r.status_code >= 400:
            raise RuntimeError(f"Erro refresh token ({r.status_code}): {r.text}")
        return r.json()

    def create_single_photo_container(self, image_url: str, caption: str) -> str:
        """Cria o 'container' de mídia (step 1 de 2 do fluxo de publish)."""
        resp = self._post(f"{self.ig_id}/media", {
            "image_url": image_url,
            "caption": caption,
        })
        return resp["id"]

    def create_carousel_container(self, image_urls: list[str], caption: str) -> str:
        """Cria carrossel de até 10 imagens."""
        children = []
        for url in image_urls:
            r = self._post(f"{self.ig_id}/media", {"image_url": url, "is_carousel_item": "true"})
            children.append(r["id"])
        resp = self._post(f"{self.ig_id}/media", {
            "media_type": "CAROUSEL",
            "caption": caption,
            "children": ",".join(children),
        })
        return resp["id"]

    def wait_container_ready(self, container_id: str, timeout: int = 300) -> None:
        """Aguarda Meta processar o container antes de publicar.
        Timeout 5min cobre vídeos grandes e momentos de fila."""
        deadline = time.time() + timeout
        while time.time() < deadline:
            r = self._get(container_id, {"fields": "status_code"})
            status = r.get("status_code")
            if status == "FINISHED":
                return
            if status in ("ERROR", "EXPIRED"):
                raise RuntimeError(f"Container falhou: {r}")
            time.sleep(4)
        raise TimeoutError(f"Container {container_id} não ficou pronto em {timeout}s")

    def publish(self, container_id: str, max_retries: int = 6, initial_wait: int = 15) -> str:
        """Publica o container (step 2 de 2).
        Retry automático pro erro 2207027 ('Media is not ready') que acontece
        quando o status FINISHED foi retornado cedo demais pela Meta.
        """
        # Pausa extra antes do 1º publish — Meta às vezes reporta FINISHED antes de estar 100% pronto
        time.sleep(initial_wait)

        last_err = None
        for attempt in range(1, max_retries + 1):
            try:
                resp = self._post(f"{self.ig_id}/media_publish", {"creation_id": container_id})
                return resp["id"]
            except RuntimeError as e:
                last_err = e
                msg = str(e)
                # Detecta erro específico de "mídia não pronta" (9007 / 2207027)
                if "2207027" in msg or "not ready" in msg.lower():
                    wait = 10 * attempt  # backoff: 10s, 20s, 30s, 40s, 50s, 60s
                    print(f"  ⏳ Tentativa {attempt}/{max_retries}: media not ready. "
                          f"Aguardando {wait}s antes de tentar de novo...")
                    time.sleep(wait)
                    continue
                # Outros erros: re-levanta imediatamente
                raise
        raise RuntimeError(f"Publish falhou após {max_retries} tentativas. Último erro: {last_err}")


# ------------------------------------------------------------ post pipeline --
def build_caption(post: dict) -> str:
    """Monta a legenda final: texto + CTA + hashtags (separados por linhas)."""
    partes = [
        str(post.get("Legenda Completa") or "").strip(),
        "",
        f"✨ {post.get('CTA') or ''}".strip(),
        "",
        str(post.get("Hashtags") or "").strip(),
    ]
    return "\n".join([p for p in partes if p is not None]).strip()


def publish_post(post: dict, cfg: dict, dry_run: bool = False) -> None:
    dia = post["Dia"]
    caption = build_caption(post)
    filename = creative_filename_for(post)

    if not filename:
        print(f"⚠  Dia {dia}: sem criativo mapeado em CREATIVES_DIR — pulando publicação automática.")
        print(f"   Legenda está pronta em '{cfg['CALENDAR_XLSX']}'. Publique manualmente.")
        return

    image_url = f"{cfg['MEDIA_BASE_URL'].rstrip('/')}/{filename}"
    print(f"→ Dia {dia} ({post['Data (2026)']}, {post['Horario']})")
    print(f"  Título: {post['Titulo / Gancho']}")
    print(f"  URL da imagem: {image_url}")
    print(f"  Caption: {len(caption)} caracteres")

    if dry_run:
        print("  [DRY-RUN] não chamando API da Meta. Caption preview:")
        print("  " + caption.replace("\n", "\n  ")[:400] + "...")
        return

    ig = IGClient(cfg["IG_USER_ID"], cfg["META_ACCESS_TOKEN"])
    print("  • Criando container...")
    container_id = ig.create_single_photo_container(image_url, caption)
    print(f"  • Container criado: {container_id}")
    print("  • Aguardando processamento...")
    ig.wait_container_ready(container_id)
    print("  • Publicando...")
    media_id = ig.publish(container_id)
    print(f"  ✓ PUBLICADO! Media ID: {media_id}")
    _append_log(cfg, {"dia": dia, "media_id": media_id, "publicado_em": datetime.utcnow().isoformat()})


def _append_log(cfg: dict, entry: dict):
    log_path = Path(__file__).parent / "publish_log.jsonl"
    with log_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


# ------------------------------------------------------------------- cli --
def main():
    parser = argparse.ArgumentParser(description="Instagram Publisher DEVOPSRAIZ")
    sub = parser.add_subparsers(dest="cmd", required=True)

    sub.add_parser("plan", help="Mostra os próximos 7 posts do calendário")
    sub.add_parser("check", help="Valida token Meta + conta Instagram")

    pub = sub.add_parser("publish", help="Publica post do calendário")
    pub.add_argument("--day", type=int, help="Número do dia (1-30)")
    pub.add_argument("--today", action="store_true", help="Publica o post do dia atual")

    dry = sub.add_parser("dry-run", help="Simula publicação sem chamar API")
    dry.add_argument("--day", type=int, required=True)

    args = parser.parse_args()
    cfg = load_env(Path(__file__).parent / ".env")

    xlsx_path = cfg.get("CALENDAR_XLSX", "../calendario-editorial-30-dias.xlsx")
    xlsx_abs = (Path(__file__).parent / xlsx_path).resolve()
    if not xlsx_abs.exists():
        sys.exit(f"Calendário não encontrado em {xlsx_abs}")

    posts = load_calendar(str(xlsx_abs))

    if args.cmd == "plan":
        print(f"Próximos 7 posts do calendário ({len(posts)} no total):\n")
        today_day = _today_day_offset()
        for p in posts:
            dia = int(p["Dia"])
            if dia < today_day or dia > today_day + 6:
                continue
            print(f"  Dia {dia:>2}  {p['Data (2026)']}  {p['Horario']}  [{p['Formato']}]  {p['Titulo / Gancho']}")
        return

    if args.cmd == "check":
        if not cfg.get("META_ACCESS_TOKEN") or not cfg.get("IG_USER_ID"):
            sys.exit("META_ACCESS_TOKEN ou IG_USER_ID não configurados no .env")
        ig = IGClient(cfg["IG_USER_ID"], cfg["META_ACCESS_TOKEN"])
        info = ig.check_identity()
        print("✓ Conectado no Instagram Business:")
        print(json.dumps(info, indent=2, ensure_ascii=False))
        return

    if args.cmd in ("publish", "dry-run"):
        target_day = args.day if args.day else _today_day_offset()
        if args.cmd == "publish" and args.today:
            target_day = _today_day_offset()
        post = next((p for p in posts if int(p["Dia"]) == target_day), None)
        if not post:
            sys.exit(f"Post do dia {target_day} não encontrado no calendário.")
        publish_post(post, cfg, dry_run=(args.cmd == "dry-run"))


def _today_day_offset() -> int:
    """Calcula qual dia do calendário corresponde a hoje (início 21/04/2026)."""
    start = date(2026, 4, 21)
    today = date.today()
    return (today - start).days + 1


if __name__ == "__main__":
    main()
